import openpyxl as pyxl
import os
import datetime as dt


def excel_list(file, sheet, dir=os.getcwd()):
    cwd = os.getcwd()
    revert = False
    if dir != cwd:
        os.chdir(dir)
        revert = True
    wb = pyxl.open(file, data_only=True)
    ws = wb[sheet]

    max_r = ws.max_row + 1
    max_c = ws.max_column + 1
    full = []
    for r in range(1, max_r):
        x = []
        for c in range(1, max_c):
            cell = pyxl.utils.get_column_letter(c) + str(r)
            x.append(ws[cell].value)

        full.append(x)

    if revert == True:
        os.chdir(cwd)
    return full

def excelcreate(data, file, sheet, dir=os.getcwd()):
    cwd = os.getcwd()
    revert = False
    if dir != cwd:
        os.chdir(dir)
        revert = True
    wb = pyxl.open(file)

    ws = wb[sheet]

    if ws['A1'] != None:
        for row in ws:
            for cell in row:
                cell.value = None
    r = 1
    for row in data:
        c = 1
        for field in row:
            if field == '#REF!' or field == '#VALUE!':
                field = None
            ws.cell(r, c).value = field
            c += 1
        r += 1

    wb.save(file)
    if revert == True:
        os.chdir(cwd)

def multiexcelwrite(file, sheetdatadic, dir=os.getcwd()):
    dic = sheetdatadic
    cwd = os.getcwd()
    revert = False
    if dir != cwd:
        os.chdir(dir)
        revert = True
    wb = pyxl.open(file)

    for sheet in list(dic.keys()):
        ws = wb[sheet]
        data = dic[sheet]
        if ws['A1'] != None:
            for row in ws:
                for cell in row:
                    cell.value = None
        r = 1
        for row in data:
            c = 1
            for field in row:
                if field == '#REF!' or field == '#VALUE!':
                    field = None
                ws.cell(r, c).value = field
                c += 1
            r += 1

    wb.save(file)
    if revert == True:
        os.chdir(cwd)


def dateconvert(data):
    for row in data:
        r_i = data.index(row)
        for x in row:
            x_i = row.index(x)
            if isinstance(x, dt.datetime):
                data[r_i][x_i] = x.date()
    return data

def project(case, act, crt):
    ans = str(case) + ' - ' + act + ' - ' + crt
    return ans

def Taskdetail():
    """RETURNS TASK NAME CHANGES IN DIC== {'Case Record Type': {'Sales Force': ['Netsuite', 'Insert Before']}}"""

    list = excel_list('Milestone Field Mapping.xlsx', 'Milestone Names')
    crt_dic = {}
    for row in list:

        if row[0] not in crt_dic.keys():
            task_dic = {}
        else:
            task_dic = crt_dic[row[0]]

        task_dic[row[1]] = [row[3], row[4]]

        crt_dic[row[0]] = task_dic

    return crt_dic



def task_len(start, end):
    if start != None:
        if end != None:
            dif = (end-start).days
        else:
            dif = (dt.datetime.today()-start).days
    else:
        dif = 0
    return dif

def binaryconvert(val):
    if val == 1:
        val = True
    elif val == 0:
        val = False

    return val

def unit_dic(units):
    dic = {}
    # dic['Headers'] = units[0]
    for row in units[1:]:
        dic[row[0]] = row[2:]

    return dic

def binloc(list):
    dic = {}
    for row in list:
        sf = row[0]
        dic[sf] = [row[1], row[2], row[4], row[5]]

    return dic



if __name__ == '__main__':
    print(excel_list('Milestone Field Mapping.xlsx', 'COMM'))